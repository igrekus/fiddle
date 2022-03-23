import uuid

import openpyxl

# обозначение
lib_C_t1 = {'ДИ': '1', 'ДИВ': '5', 'ДВ': '4', 'ДА': '2'}  # Обозначение/Наименование
lib_X_t3 = {'0,25': '', '0,5': '-1', '1,5': '-2'}  # Класс точности


# спецификация
# поз. 1
lib_G_t2_p1 = {'1': 'ПС2-DAT09C04-01', '2': 'ПС2-DAT09C04-01',
               '7': 'ПС2-DAT09C04-01', '8': 'ПС2-DAT09C04-01',
               '4': 'ПС2-DAT41C01-01'}
lib_G_t2_p1_EXI = {'1': 'ПС2-DAT09C04-02', '2': 'ПС2-DAT09C04-02',
                   '7': 'ПС2-DAT09C04-02', '8': 'ПС2-DAT09C04-02',
                   '4': 'ПС2-DAT41C01-02'}
# поз. 2
lib_Y_t2_p2 = {'': 'КУВФ.301152.320.1175-01', 'EXI': 'КУВФ.301152.320.1175-02'}

# поз. 5
lib_G_t4_p5 = {'1': 'КУВФ.713541.430.2054 (Штуцер М20х1,5)',
               '2': '',
               '4': 'КУВФ.713567.430.2771 (Штуцер М24х1,5 торц.)',
               '7': 'КУВФ.713567.430.2183 (Штуцер G1/2 B)',
               '8': 'КУВФ.713567.430.8174 (Штуцер G1/2 B)'}
# поз. 12
lib_G_t4_p12 = {'1': '',
                '2': '',
                '4': 'Кольцо 025-028-19 ГОСТ 9833',
                '7': '',
                '8': 'Кольцо 025-028-19 ГОСТ 9833'}

# Комплекты
lib_G_t4_pack = {'1': 'КУВФ.407975.390.408-01 (Упаковка 408)',
                 '2': 'КУВФ.407975.390.408 (Упаковка 408)',
                 '4': 'КУВФ.407975.390.408-06 (Упаковка 408)',
                 '7': 'КУВФ.407975.390.408-01 (Упаковка 408)',
                 '8': 'КУВФ.407975.390.408 (Упаковка 408)'}

# поз. 8
lib_G1478_t1_p8_DI = {'0,01': '4480-19-0,01',
                      '0,016': '4480-19-0,035',
                      '0,025': '4480-19-0,035',
                      '0,04': '4480-19-0,1',
                      '0,06': '4480-19-0,1',
                      '0,1': '4480-19-0,1',
                      '0,16': '4480-19-0,25',
                      '0,25': '4480-19-0,25',
                      '0,4': '4480-19-0,5',
                      '0,6': '4480-19-1,0',
                      '1': '4480-19-1,0',
                      '1,0': '4480-19-1,0',
                      '1,6': '4480-19-3,0',
                      '2,5': '4480-19-3,0',
                      '4': '4480-19-10,0',
                      '4,0': '4480-19-10,0'}
lib_G2_t1_p8_DI = {'0,01': '4480-G1/2-0,01',
                   '0,016': '4480-G1/2-0,035',
                   '0,025': '4480-G1/2-0,035',
                   '0,04': '4480-G1/2-0,1',
                   '0,06': '4480-G1/2-0,1',
                   '0,1': '4480-G1/2-0,1',
                   '0,16': '4480-G1/2-0,25',
                   '0,25': '4480-G1/2-0,25',
                   '0,4': '4480-G1/2-0,5',
                   '0,6': '4480-G1/2-1,0',
                   '1': '4480-G1/2-1,0',
                   '1,0': '4480-G1/2-1,0',
                   '1,6': '4480-G1/2-3,0',
                   '2,5': '4480-G1/2-3,0',
                   '4': '4480-G1/2-10,0',
                   '4,0': '4480-G1/2-10,0'}
lib_G1478_t1_p8_DIV = {'0,0125': '4480-19-0,01',
                       '0,02': '4480-19-0,035',
                       '0,03': '4480-19-0,035',
                       '0,05': '4480-19-0,1',
                       '0,08': '4480-19-0,1',
                       '0,1': '4480-19-0,1',
                       '0,15': '4480-19-0,25',
                       '0,3': '4480-19-0,5',
                       '0,5': '4480-19-0,5',
                       '0,9': '4480-19-1,0',
                       '1,5': '4480-19-3,0',
                       '2,4': '4480-19-3,0'}
lib_G2_t1_p8_DIV = {'0,0125': '4480-G1/2-0,01',
                    '0,02': '4480-G1/2-0,035',
                    '0,03': '4480-G1/2-0,035',
                    '0,05': '4480-G1/2-0,1',
                    '0,08': '4480-G1/2-0,1',
                    '0,1': '4480-G1/2-0,1',
                    '0,15': '4480-G1/2-0,25',
                    '0,3': '4480-G1/2-0,5',
                    '0,5': '4480-G1/2-0,5',
                    '0,9': '4480-G1/2-1,0',
                    '1,5': '4480-G1/2-3,0',
                    '2,4': '4480-G1/2-3,0'}
lib_G1478_t1_p8_DV = {'0,01': '4480-19-0,01',
                      '0,016': '4480-19-0,035',
                      '0,025': '4480-19-0,035',
                      '0,04': '4480-19-0,1',
                      '0,06': '4480-19-0,1',
                      '0,1': '4480-19-0,1'}
lib_G2_t1_p8_DV = {'0,01': '4480-G1/2-0,01',
                   '0,016': '4480-G1/2-0,035',
                   '0,025': '4480-G1/2-0,035',
                   '0,04': '4480-G1/2-0,1',
                   '0,06': '4480-G1/2-0,1',
                   '0,1': '4480-G1/2-0,1'}
lib_G1478_t1_p8_DA = {'0,1': '4431-19-0,1',
                      '0,16': '4431-19-0,25',
                      '0,25': '4431-19-0,25',
                      '0,4': '4431-19-0,5',
                      '0,6': '4431-19-1,0',
                      '1': '4431-19-1,0',
                      '1,0': '4431-19-1,0',
                      '1,6': '4431-19-3,0'}


def read():
    wb = openpyxl.load_workbook(filename='base.xlsx')
    sheet = wb.worksheets[0]
    indexes = []

    for row in sheet.values:
        for cell in row:
            indexes.append(cell)

    return indexes


def save(indexes):
    spec_name = 'test.xlsx'

    # first sheet
    n = 0
    for index in indexes[:5]:
        wb = openpyxl.load_workbook(filename=spec_name)
        ws = wb.worksheets[0]
        index = str(indexes[n]).split('-')
        for i in index:
            ws['A2'].value = n
            if index[1].startswith('ДИ') and not index[1].startswith('ДИВ'):
                if len(index) == 4:
                    ws['B2'].value = 'КУВФ.406233.101-' + index[1][2:] + \
                                     "-61" + index[2][1] + '1' + lib_X_t3[index[3]] + \
                                     ' (ДУП-' + indexes[n] + ')'
                if len(index) == 5:
                    ws['B2'].value = 'КУВФ.406233.101-' + index[1][2:] + \
                                     '-61' + index[2][1] + '1EXI' + lib_X_t3[index[3]] + \
                                     ' (ДУП-' + indexes[n] + ')'
            elif index[1].startswith('ДИВ'):
                if len(index) == 4:
                    ws['B2'].value = 'КУВФ.406233.105-' + index[1][3:] + \
                                     '-61' + index[2][1] + '1' + lib_X_t3[index[3]] + \
                                     ' (ДУП-' + indexes[n] + ')'
                if len(index) == 5:
                    ws['B2'].value = 'КУВФ.406233.105-' + index[1][3:] + \
                                     '-61' + index[2][1] + '1EXI' + lib_X_t3[index[3]] + \
                                     ' (ДУП-' + indexes[n] + ')'
            elif index[1].startswith('ДВ'):
                if len(index) == 4:
                    ws['B2'].value = 'КУВФ.406233.104-' + index[1][2:] + \
                                     '-61' + index[2][1] + '1' + lib_X_t3[index[3]] + \
                                     ' (ДУП-' + indexes[n] + ')'
                if len(index) == 5:
                    ws['B2'].value = 'КУВФ.406233.104-' + index[1][2:] + \
                                     '-61' + index[2][1] + '1EXI' + lib_X_t3[index[3]] + \
                                     ' (ДУП-' + indexes[n] + ')'
            elif index[1].startswith('ДА'):
                if len(index) == 4:
                    ws['B2'].value = 'КУВФ.406233.102-' + index[1][2:] + \
                                     '-61' + index[2][1] + '1' + lib_X_t3[index[3]] + \
                                     ' (ДУП-' + indexes[n] + ')'
                if len(index) == 5:
                    ws['B2'].value = 'КУВФ.406233.102-' + index[1][2:] + \
                                     '-61' + index[2][1] + '1EXI' + lib_X_t3[index[3]] + \
                                     ' (ДУП-' + indexes[n] + ')'

        # second sheet, constant data
        ws = wb.worksheets[1]

        ws['A2'].value = '1'
        ws['A3'].value = '2'
        ws['A4'].value = '3'
        ws['A5'].value = '5'
        ws['A6'].value = '8'
        ws['A7'].value = '9'
        ws['A8'].value = '10'
        ws['A9'].value = '12'
        ws['A10'].value = '14'
        ws['A11'].value = '16'

        ws['B2'].value = 'Узел'
        ws['B3'].value = 'Узел'
        ws['B4'].value = 'Деталь'
        ws['B5'].value = 'Деталь'
        ws['B6'].value = 'Материал'
        ws['B7'].value = 'Материал'
        ws['B8'].value = 'Материал'
        ws['B9'].value = 'Материал'
        ws['B10'].value = 'Материал'
        ws['B11'].value = 'Материал'
        ws['B12'].value = 'Набор'

        ws['D2'].value = '1,000'
        ws['D3'].value = '1,000'
        ws['D4'].value = '1,000'
        ws['D5'].value = '1,000'
        ws['D6'].value = '1,000'
        ws['D7'].value = '1,000'
        ws['D8'].value = '1,000'
        ws['D9'].value = '1,000'
        ws['D10'].value = '1,000'
        ws['D11'].value = '1,000'
        ws['D12'].value = '1,000'

        ws['C4'].value = 'КУВФ.715141.430.2085 (Корпус)'
        ws['C7'].value = 'Силиконовый компаунд Wacker SilGel 612A'
        ws['C8'].value = 'Силиконовый компаунд Wacker SilGel 612B'
        ws['C10'].value = 'Клей "Супер Момент" секундный'
        ws['C11'].value = 'Припой безотмывочный REL61 GlowCore, 0,5мм, AIM'

        # second sheet variable data
        ind = str(indexes[n]).split('-')
        for j in ind:
            # поз. 1
            if len(ind) == 4:
                if ind[2][1] == '1' or '2' or '7' or '8':
                    ws['C2'].value = lib_G_t2_p1[ind[2][1]]
                elif ind[2][1] == '4':
                    ws['C2'].value = lib_G_t2_p1[ind[2][1]]
            if len(ind) == 5:
                if ind[2][1] == '1' or '2' or '7' or '8':
                    ws['C2'].value = lib_G_t2_p1_EXI[ind[2][1]]
                elif ind[2][1] == '4':
                    ws['C2'].value = lib_G_t2_p1_EXI[ind[2][1]]

            # поз. 2
            if len(ind) == 4:
                ws['C3'].value = lib_Y_t2_p2['']
            if len(ind) == 5:
                ws['C3'].value = lib_Y_t2_p2['EXI']

            # поз. 5
            if ind[2][1] == '1':
                ws['C5'].value = lib_G_t4_p5[ind[2][1]]
            elif ind[2][1] == '2':
                ws['C5'].value = lib_G_t4_p5[ind[2][1]]
                ws['D5'].value = ''
                ws['B5'].value = ''
                ws['A5'].value = ''
            elif ind[2][1] == '4':
                ws['C5'].value = lib_G_t4_p5[ind[2][1]]
            elif ind[2][1] == '7':
                ws['C5'].value = lib_G_t4_p5[ind[2][1]]
            elif ind[2][1] == '8':
                ws['C5'].value = lib_G_t4_p5[ind[2][1]]

            # поз. 12
            if ind[2][1] == '1':
                ws['C9'].value = lib_G_t4_p12[ind[2][1]]
                ws['D9'].value = ''
                ws['B9'].value = ''
                ws['A9'].value = ''
            elif ind[2][1] == '2':
                ws['C9'].value = lib_G_t4_p12[ind[2][1]]
                ws['D9'].value = ''
                ws['B9'].value = ''
                ws['A9'].value = ''
            elif ind[2][1] == '4':
                ws['C9'].value = lib_G_t4_p12[ind[2][1]]
            elif ind[2][1] == '7':
                ws['C9'].value = lib_G_t4_p12[ind[2][1]]
                ws['D9'].value = ''
                ws['B9'].value = ''
                ws['A9'].value = ''
            elif ind[2][1] == '8':
                ws['C9'].value = lib_G_t4_p12[ind[2][1]]

            # Комплекты
            if ind[2][1] == '1':
                ws['C12'].value = lib_G_t4_pack[ind[2][1]]
            elif ind[2][1] == '2':
                ws['C12'].value = lib_G_t4_pack[ind[2][1]]
            elif ind[2][1] == '4':
                ws['C12'].value = lib_G_t4_pack[ind[2][1]]
            elif ind[2][1] == '7':
                ws['C12'].value = lib_G_t4_pack[ind[2][1]]
            elif ind[2][1] == '8':
                ws['C12'].value = lib_G_t4_pack[ind[2][1]]

            # поз. 8
            if ind[1].startswith('ДИ') and not ind[1].startswith('ДИВ'):
                if ind[2][1] == '1' or '4' or '7' or '8':
                    ws['C6'].value = lib_G1478_t1_p8_DI[ind[1][2:]]
                elif ind[2][1] == '2':
                    ws['C6'].value = lib_G2_t1_p8_DI[ind[1][2:]]
            if ind[1].startswith('ДИВ'):
                if ind[2][1] == '1' or '4' or '7' or '8':
                    ws['C6'].value = lib_G1478_t1_p8_DIV[ind[1][3:]]
                elif ind[2][1] == '2':
                    ws['C6'].value = lib_G2_t1_p8_DIV[ind[1][3:]]
            if ind[1].startswith('ДВ'):
                if ind[2][1] == '1' or '4' or '7' or '8':
                    ws['C6'].value = lib_G1478_t1_p8_DV[ind[1][2:]]
                elif ind[2][1] == '2':
                    ws['C6'].value = lib_G2_t1_p8_DV[ind[1][2:]]
            if ind[1].startswith('ДА'):
                if ind[2][1] == '1' or '4' or '7' or '8':
                    ws['C6'].value = lib_G1478_t1_p8_DA[ind[1][2:]]
                else:
                    ws['A6'].value = ''
                    ws['B6'].value = ''
                    ws['C6'].value = ''
                    ws['D6'].value = ''

        wb.save(spec_name + uuid.uuid4().hex + '.xlsx')
        n += 1


if __name__ == '__main__':
    idx = read()
    save(idx)
